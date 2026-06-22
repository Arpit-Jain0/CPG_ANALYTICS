#!/usr/bin/env python3
"""
scripts/generate_data.py

Deterministic synthetic-data generator for the CPG Analytics Platform.
Writes realistic, signal-bearing .xlsx workbooks into the input folders.

Usage
-----
    python scripts/generate_data.py              # seed=42
    python scripts/generate_data.py --seed 7
    python scripts/generate_data.py --root /alt/repo/root
"""

from __future__ import annotations

import argparse
from datetime import date, datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════
DEFAULT_SEED = 42
HIST_START = date(2022, 7, 1)
HIST_END = date(2024, 6, 30)
CAL_END = date(2025, 12, 31)  # calendar extends past the data window


# ═══════════════════════════════════════════════════════════════════════════════
# MASTER REFERENCE DATA
# ═══════════════════════════════════════════════════════════════════════════════
PRODUCT_ROWS = [
    # (sku, category, brand, package_size, list_price, launch_date)
    ("SKU001", "Beverages", "RefreshCo", "12-pack 355ml", 8.99, date(2018, 3, 1)),
    ("SKU002", "Beverages", "NaturePlus", "1L bottle", 3.49, date(2019, 6, 15)),
    ("SKU003", "Snacks", "CrunchMaster", "150g bag", 2.99, date(2017, 1, 10)),
    ("SKU004", "Snacks", "GrainGoods", "400g box", 5.49, date(2020, 4, 1)),
    ("SKU005", "Personal Care", "GlowBrand", "250ml shampoo", 6.99, date(2016, 9, 1)),
    ("SKU006", "Personal Care", "GlowBrand", "250ml conditioner", 7.49, date(2016, 9, 1)),
    ("SKU007", "Household", "CleanPro", "1kg detergent", 9.99, date(2015, 2, 28)),
    ("SKU008", "Household", "CleanPro", "500ml spray", 4.99, date(2015, 2, 28)),
    ("SKU009", "Frozen Foods", "FrostBite", "500g pizza", 7.99, date(2021, 11, 1)),
    ("SKU010", "Frozen Foods", "FrostBite", "600g nuggets", 6.49, date(2021, 11, 1)),
    ("SKU011", "Beverages", "AquaPure", "6-pack 500ml", 4.99, date(2019, 3, 15)),
    ("SKU012", "Snacks", "SweetTreat", "200g cookies", 3.99, date(2022, 1, 5)),
]

# Average Poisson rate: transactions per (store-SKU-day) for clean signal
CATEGORY_BASE_RATE: dict[str, float] = {
    "Beverages": 0.55,
    "Snacks": 0.50,
    "Personal Care": 0.22,
    "Household": 0.18,
    "Frozen Foods": 0.30,
}

REGION_ROWS = [
    # (region, population, median_income_band, climate_zone)
    ("NORTHEAST", 15_000_000, "HIGH", "Temperate"),
    ("SOUTHEAST", 12_000_000, "MEDIUM", "Subtropical"),
    ("MIDWEST", 10_000_000, "MEDIUM", "Continental"),
    ("WEST", 18_000_000, "HIGH", "Mediterranean"),
]

STORE_ROWS = [
    # Physical POS stores
    ("STORE001", "NORTHEAST", "New York", "SUPERMARKET"),
    ("STORE002", "NORTHEAST", "Boston", "CONVENIENCE"),
    ("STORE003", "SOUTHEAST", "Atlanta", "SUPERMARKET"),
    ("STORE004", "SOUTHEAST", "Miami", "SUPERMARKET"),
    ("STORE005", "MIDWEST", "Chicago", "SUPERMARKET"),
    ("STORE006", "MIDWEST", "Detroit", "CONVENIENCE"),
    ("STORE007", "WEST", "Los Angeles", "SUPERMARKET"),
    ("STORE008", "WEST", "Seattle", "CONVENIENCE"),
    # Online virtual locations (one per region, used as location_id in schema B)
    ("ONLINE-NE", "NORTHEAST", "Online", "ONLINE"),
    ("ONLINE-SE", "SOUTHEAST", "Online", "ONLINE"),
    ("ONLINE-MW", "MIDWEST", "Online", "ONLINE"),
    ("ONLINE-WE", "WEST", "Online", "ONLINE"),
]

POS_STORES = [r[0] for r in STORE_ROWS if r[3] != "ONLINE"]
ONLINE_LOCS = [r[0] for r in STORE_ROWS if r[3] == "ONLINE"]
STORE_TO_REGION = {r[0]: r[1] for r in STORE_ROWS}
SKU_TO_CATEGORY = {r[0]: r[1] for r in PRODUCT_ROWS}
SKU_TO_PRICE = {r[0]: r[4] for r in PRODUCT_ROWS}

PROMO_ROWS = [
    # (promo_id, category, region, start_date, end_date, discount_pct)
    ("PROMO001", "Beverages", "NORTHEAST", date(2022, 11, 25), date(2022, 12, 31), 15.0),
    ("PROMO002", "Snacks", "SOUTHEAST", date(2023, 2, 1), date(2023, 2, 28), 10.0),
    ("PROMO003", "Household", "MIDWEST", date(2023, 5, 15), date(2023, 6, 15), 20.0),
    ("PROMO004", "Beverages", "WEST", date(2023, 7, 4), date(2023, 7, 31), 12.0),
    ("PROMO005", "Frozen Foods", "NORTHEAST", date(2023, 10, 1), date(2023, 11, 30), 18.0),
    ("PROMO006", "Snacks", "MIDWEST", date(2024, 1, 15), date(2024, 2, 15), 8.0),
    ("PROMO007", "Personal Care", "WEST", date(2024, 3, 1), date(2024, 3, 31), 15.0),
    ("PROMO008", "Beverages", "SOUTHEAST", date(2024, 5, 1), date(2024, 5, 31), 10.0),
]

CAMPAIGN_ROWS = [
    # (campaign_id, category, region, channel, start_date, end_date, exposure)
    ("CAMP001", "Beverages", "NORTHEAST", "TV", date(2022, 10, 1), date(2022, 11, 30), 5_000_000),
    ("CAMP002", "Snacks", "WEST", "SOCIAL", date(2023, 1, 1), date(2023, 3, 31), 2_500_000),
    ("CAMP003", "Household", "SOUTHEAST", "PRINT", date(2023, 4, 1), date(2023, 5, 31), 1_200_000),
    (
        "CAMP004",
        "Personal Care",
        "MIDWEST",
        "DIGITAL",
        date(2023, 8, 1),
        date(2023, 10, 31),
        3_800_000,
    ),
    ("CAMP005", "Beverages", "WEST", "TV", date(2024, 1, 1), date(2024, 3, 31), 6_200_000),
    (
        "CAMP006",
        "Frozen Foods",
        "NORTHEAST",
        "SOCIAL",
        date(2024, 4, 1),
        date(2024, 6, 30),
        1_800_000,
    ),
]


# ═══════════════════════════════════════════════════════════════════════════════
# SEASONAL CALENDAR
# ═══════════════════════════════════════════════════════════════════════════════
_MONTH_TO_SEASON = {
    1: "WINTER",
    2: "WINTER",
    3: "SPRING",
    4: "SPRING",
    5: "SPRING",
    6: "SUMMER",
    7: "SUMMER",
    8: "SUMMER",
    9: "FALL",
    10: "FALL",
    11: "FALL",
    12: "WINTER",
}


def _nth_weekday(year: int, month: int, weekday: int, n: int) -> date:
    """Return the nth occurrence (1-indexed) of weekday (0=Mon) in year/month."""
    first = date(year, month, 1)
    offset = (weekday - first.weekday()) % 7
    return first + timedelta(days=offset + 7 * (n - 1))


def _last_weekday(year: int, month: int, weekday: int) -> date:
    nxt = date(year, month % 12 + 1, 1) if month < 12 else date(year + 1, 1, 1)
    last_day = nxt - timedelta(days=1)
    offset = (last_day.weekday() - weekday) % 7
    return last_day - timedelta(days=offset)


def _us_holidays(year: int) -> dict[date, str]:
    return {
        date(year, 1, 1): "New Year's Day",
        _nth_weekday(year, 1, 0, 3): "MLK Day",
        _nth_weekday(year, 2, 0, 3): "Presidents' Day",
        _last_weekday(year, 5, 0): "Memorial Day",
        date(year, 7, 4): "Independence Day",
        _nth_weekday(year, 9, 0, 1): "Labor Day",
        _nth_weekday(year, 10, 0, 2): "Columbus Day",
        date(year, 11, 11): "Veterans Day",
        _nth_weekday(year, 11, 3, 4): "Thanksgiving",
        date(year, 12, 25): "Christmas",
    }


def make_seasonal_calendar() -> pd.DataFrame:
    holidays: dict[date, str] = {}
    for y in range(HIST_START.year, CAL_END.year + 1):
        holidays.update(_us_holidays(y))

    rows, cur = [], HIST_START
    while cur <= CAL_END:
        hname = holidays.get(cur)
        rows.append(
            {
                "calendar_date": cur,
                "season": _MONTH_TO_SEASON[cur.month],
                "is_holiday": hname is not None,
                "holiday_name": hname,
            }
        )
        cur += timedelta(days=1)
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════════════
# DEMAND SIGNAL
# ═══════════════════════════════════════════════════════════════════════════════
def build_promo_set() -> set[tuple[date, str, str]]:
    """Build (date, category, region) lookup for all promo calendar days."""
    s: set[tuple[date, str, str]] = set()
    for _, cat, reg, start, end, _ in PROMO_ROWS:
        cur = start
        while cur <= end:
            s.add((cur, cat, reg))
            cur += timedelta(days=1)
    return s


def demand_mult(
    dt: date,
    category: str,
    region: str,
    promo_set: set,
    rng: np.random.Generator,
) -> float:
    """Demand multiplier = trend × weekly × yearly × promo × noise."""
    months = (dt.year - HIST_START.year) * 12 + (dt.month - HIST_START.month)
    trend = 1.0 + 0.002 * months  # 0.2 %/month uplift
    weekly = 1.30 if dt.weekday() >= 5 else 1.0  # 30 % weekend lift
    yearly = 1.35 if dt.month in (10, 11, 12) else 1.0  # Q4 bump
    promo = 1.25 if (dt, category, region) in promo_set else 1.0
    noise = float(rng.lognormal(mean=0.0, sigma=0.18))
    return trend * weekly * yearly * promo * noise


# ═══════════════════════════════════════════════════════════════════════════════
# TRANSACTION ID COUNTERS  (module-level so historical and incremental
# IDs never collide)
# ═══════════════════════════════════════════════════════════════════════════════
_txn_counter = 0
_ord_counter = 0


def _next_txn() -> str:
    global _txn_counter
    _txn_counter += 1
    return f"TXN-{_txn_counter:08d}"


def _next_ord() -> str:
    global _ord_counter
    _ord_counter += 1
    return f"ORD-{_ord_counter:08d}"


# ═══════════════════════════════════════════════════════════════════════════════
# TRANSACTION GENERATORS
# ═══════════════════════════════════════════════════════════════════════════════
def gen_pos_transactions(
    date_range: list[date],
    promo_set: set,
    rng: np.random.Generator,
) -> pd.DataFrame:
    """Generate clean POS transactions (schema A — 8 columns, has 'amount')."""
    records: list[dict] = []
    for dt in date_range:
        for store_id in POS_STORES:
            region = STORE_TO_REGION[store_id]
            for sku, category, _, _, list_price, _ in PRODUCT_ROWS:
                lam = CATEGORY_BASE_RATE[category] * demand_mult(
                    dt, category, region, promo_set, rng
                )
                for _ in range(int(rng.poisson(lam))):
                    hour = int(rng.integers(8, 21))
                    minute = int(rng.integers(0, 60))
                    qty = max(1, int(rng.integers(1, 6)))
                    unit_price = round(float(list_price * rng.normal(1.0, 0.03)), 2)
                    records.append(
                        {
                            "transaction_id": _next_txn(),
                            "ts": datetime(dt.year, dt.month, dt.day, hour, minute),
                            "store_id": store_id,
                            "sku": sku,
                            "qty": qty,
                            "unit_price": unit_price,
                            "amount": round(qty * unit_price, 2),
                            "currency": "USD",
                        }
                    )
    return pd.DataFrame(records)


def gen_online_transactions(
    date_range: list[date],
    promo_set: set,
    rng: np.random.Generator,
) -> pd.DataFrame:
    """Generate clean online transactions (schema B — drifted, no amount column)."""
    records: list[dict] = []
    for dt in date_range:
        for loc_id in ONLINE_LOCS:
            region = STORE_TO_REGION[loc_id]
            for sku, category, _, _, list_price, _ in PRODUCT_ROWS:
                # online volume ~45 % of per-POS-store rate
                lam = (
                    CATEGORY_BASE_RATE[category]
                    * 0.45
                    * demand_mult(dt, category, region, promo_set, rng)
                )
                for _ in range(int(rng.poisson(lam))):
                    hour = int(rng.integers(0, 24))
                    minute = int(rng.integers(0, 60))
                    units = max(1, int(rng.integers(1, 8)))
                    price_per_unit = round(float(list_price * rng.normal(1.0, 0.04)), 2)
                    records.append(
                        {
                            "order_id": _next_ord(),
                            "order_datetime": datetime(dt.year, dt.month, dt.day, hour, minute),
                            "location_id": loc_id,
                            "product_sku": sku,
                            "units": units,
                            "price_per_unit": price_per_unit,
                            # intentionally no amount/revenue column — schema drift
                            "currency": "USD",
                        }
                    )
    return pd.DataFrame(records)


# ═══════════════════════════════════════════════════════════════════════════════
# DQ ISSUE INJECTION
# ═══════════════════════════════════════════════════════════════════════════════
_MIXED_FMT_FNS = [
    lambda ts: ts.strftime("%m/%d/%Y %H:%M"),  # US slash
    lambda ts: ts.strftime("%d-%b-%Y %H:%M"),  # text month
    lambda ts: ts.strftime("%Y/%m/%d %H:%M:%S"),  # slash-ISO
]


def inject_dq(
    df: pd.DataFrame,
    schema: str,  # "A" = POS, "B" = online
    rng: np.random.Generator,
    dq_counts: dict[str, int],
) -> pd.DataFrame:
    """
    Inject ~8.8 % DQ issues into a copy of df.
    Index sets for each issue are non-overlapping so NULL+NULL rows don't appear
    (which would make them unrepariable).
    """
    df = df.copy()
    n = len(df)
    perm = rng.permutation(n)

    rates: dict[str, float] = {
        "NULL_UNIT_PRICE": 0.020,
        "NULL_AMOUNT": 0.020 if schema == "A" else 0.0,
        "MIXED_DATE_FMT": 0.015,
        "EUR_CURRENCY": 0.015,
        "ZERO_NEG_QTY": 0.008,
        "UNKNOWN_STORE_ID": 0.005,
        "UNKNOWN_SKU": 0.005,
    }

    ptr = 0
    buckets: dict[str, np.ndarray] = {}
    for issue, rate in rates.items():
        cnt = int(n * rate)
        buckets[issue] = perm[ptr : ptr + cnt]
        ptr += cnt
        dq_counts[issue] = dq_counts.get(issue, 0) + cnt

    # Column aliases per schema
    price_col = "unit_price" if schema == "A" else "price_per_unit"
    qty_col = "qty" if schema == "A" else "units"
    ts_col = "ts" if schema == "A" else "order_datetime"
    store_col = "store_id" if schema == "A" else "location_id"
    sku_col = "sku" if schema == "A" else "product_sku"

    # 1. Null unit_price  (repairable: amount / qty)
    df.loc[buckets["NULL_UNIT_PRICE"], price_col] = np.nan

    # 2. Null amount  (repairable: qty * unit_price)  — schema A only
    if schema == "A":
        df.loc[buckets["NULL_AMOUNT"], "amount"] = np.nan

    # 3. Mixed date formats — convert datetime → string in a non-ISO format
    for i in buckets["MIXED_DATE_FMT"]:
        val = df.at[i, ts_col]
        if isinstance(val, datetime):
            fmt_fn = _MIXED_FMT_FNS[int(i) % len(_MIXED_FMT_FNS)]
            df.at[i, ts_col] = fmt_fn(val)

    # 4. EUR currency  (normalise to USD during ingestion)
    df.loc[buckets["EUR_CURRENCY"], "currency"] = "EUR"

    # 5. Zero / negative quantity  (reject)
    neg = buckets["ZERO_NEG_QTY"]
    half = max(1, len(neg) // 2)
    df.loc[neg[:half], qty_col] = 0
    df.loc[neg[half:], qty_col] = -int(rng.integers(1, 4))

    # 6. Unknown store_id
    df.loc[buckets["UNKNOWN_STORE_ID"], store_col] = "STORE999"

    # 7. Unknown SKU
    df.loc[buckets["UNKNOWN_SKU"], sku_col] = "SKU999"

    return df


# ═══════════════════════════════════════════════════════════════════════════════
# SECONDARY FEEDS
# ═══════════════════════════════════════════════════════════════════════════════
def make_promo_df() -> pd.DataFrame:
    cols = ["promo_id", "category", "region", "start_date", "end_date", "discount_pct"]
    return pd.DataFrame(PROMO_ROWS, columns=cols)


def make_campaign_df() -> pd.DataFrame:
    cols = ["campaign_id", "category", "region", "channel", "start_date", "end_date", "exposure"]
    return pd.DataFrame(CAMPAIGN_ROWS, columns=cols)


def make_competitor_prices(rng: np.random.Generator) -> pd.DataFrame:
    """Monthly competitor-price observations (one row per category × region × month)."""
    categories = list({r[1] for r in PRODUCT_ROWS})
    regions = [r[0] for r in REGION_ROWS]
    rows: list[dict] = []
    cur = date(HIST_START.year, HIST_START.month, 1)
    while cur <= HIST_END:
        for cat in categories:
            avg_price = float(np.mean([SKU_TO_PRICE[s] for s, c, *_ in PRODUCT_ROWS if c == cat]))
            for reg in regions:
                rows.append(
                    {
                        "obs_date": cur,
                        "category": cat,
                        "region": reg,
                        "competitor_price": round(avg_price * float(rng.normal(1.05, 0.08)), 2),
                    }
                )
        cur = date(cur.year + (cur.month == 12), cur.month % 12 + 1, 1)
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════════════
# SCD-2 PRODUCT CHANGE (for incremental batch 2)
# ═══════════════════════════════════════════════════════════════════════════════
def make_scd2_updates() -> pd.DataFrame:
    """
    Product-attribute changes effective 2024-07-08.
    Ingestion must detect these rows and create new Type-2 SCD records.
    """
    return pd.DataFrame(
        [
            {
                "sku": "SKU001",
                "attribute": "list_price",
                "old_value": "8.99",
                "new_value": "9.49",
                "effective_date": date(2024, 7, 8),
                "change_reason": "Annual price revision",
            },
            {
                "sku": "SKU007",
                "attribute": "package_size",
                "old_value": "1kg detergent",
                "new_value": "1.2kg detergent",
                "effective_date": date(2024, 7, 8),
                "change_reason": "Reformulation — larger pack",
            },
        ]
    )


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL WRITING HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def _df_to_ws(ws, df: pd.DataFrame) -> None:
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)


def write_multisheet(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    """Write one workbook with multiple named sheets in insertion order."""
    wb = Workbook()
    wb.remove(wb.active)  # delete the default blank sheet
    for name, df in sheets.items():
        ws = wb.create_sheet(title=name)
        _df_to_ws(ws, df)
    wb.save(path)


def write_single_sheet(
    path: Path,
    df: pd.DataFrame,
    sheet_name: str = "Sheet1",
    title_row: str | None = None,
    ghost_sheet: bool = False,
) -> None:
    """
    Write a single-sheet workbook.
    title_row   → inserts a metadata/title row ABOVE the header (hard-path for readers).
    ghost_sheet → adds an empty 'Notes' sheet alongside the data sheet.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    if title_row is not None:
        ws.append([title_row])  # row 1 = opaque metadata
    _df_to_ws(ws, df)
    if ghost_sheet:
        wb.create_sheet(title="Notes")  # empty stale sheet
    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════════════
# INCREMENTAL BATCH HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def _make_late_rows(
    base_df: pd.DataFrame,
    ts_col: str,
    id_col: str,
    id_gen,
    n: int,
    late_before: date,
    rng: np.random.Generator,
) -> pd.DataFrame:
    rows = base_df.sample(min(n, len(base_df)), random_state=int(rng.integers(9999))).copy()
    for idx in rows.index:
        days_back = int(rng.integers(2, 8))
        late_dt = late_before - timedelta(days=days_back)
        orig = rows.at[idx, ts_col]
        if isinstance(orig, datetime):
            rows.at[idx, ts_col] = datetime(
                late_dt.year, late_dt.month, late_dt.day, orig.hour, orig.minute
            )
    rows[id_col] = [id_gen() for _ in range(len(rows))]
    return rows


def gen_incr_pos_batch(
    week_start: date,
    promo_set: set,
    rng: np.random.Generator,
    dup_ids: list[str],
    n_late: int,
    late_before: date,
) -> tuple[pd.DataFrame, dict[str, int]]:
    dq: dict[str, int] = {}
    dates = [week_start + timedelta(days=i) for i in range(7)]
    df = gen_pos_transactions(dates, promo_set, rng)
    df = inject_dq(df, "A", rng, dq)

    extras = []

    # Duplicate rows — retry double-send
    if dup_ids:
        n_dup = min(len(dup_ids), len(df))
        dups = df.sample(n_dup, random_state=int(rng.integers(9999))).copy()
        dups["transaction_id"] = dup_ids[:n_dup]
        extras.append(dups)
        dq["DUPLICATE_ID"] = n_dup

    # Late-arriving rows
    if n_late > 0:
        late = _make_late_rows(df, "ts", "transaction_id", _next_txn, n_late, late_before, rng)
        extras.append(late)
        dq["LATE_ARRIVING"] = len(late)

    if extras:
        df = pd.concat([df, *extras], ignore_index=True)
    return df, dq


def gen_incr_online_batch(
    week_start: date,
    promo_set: set,
    rng: np.random.Generator,
    dup_ids: list[str],
    n_late: int,
    late_before: date,
) -> tuple[pd.DataFrame, dict[str, int]]:
    dq: dict[str, int] = {}
    dates = [week_start + timedelta(days=i) for i in range(7)]
    df = gen_online_transactions(dates, promo_set, rng)
    df = inject_dq(df, "B", rng, dq)

    extras = []

    if dup_ids:
        n_dup = min(len(dup_ids), len(df))
        dups = df.sample(n_dup, random_state=int(rng.integers(9999))).copy()
        dups["order_id"] = dup_ids[:n_dup]
        extras.append(dups)
        dq["DUPLICATE_ID"] = n_dup

    if n_late > 0:
        late = _make_late_rows(
            df, "order_datetime", "order_id", _next_ord, n_late, late_before, rng
        )
        extras.append(late)
        dq["LATE_ARRIVING"] = len(late)

    if extras:
        df = pd.concat([df, *extras], ignore_index=True)
    return df, dq


# ═══════════════════════════════════════════════════════════════════════════════
# SUMMARY PRINTER
# ═══════════════════════════════════════════════════════════════════════════════
def print_summary(stats: list[dict], seed: int) -> None:
    W = 68
    print("\n" + "═" * W)
    print(f"  CPG Analytics — Synthetic Data Generator  (seed={seed})")
    print("═" * W)

    section = None
    for s in stats:
        if s.get("section") != section:
            section = s["section"]
            print(f"\n{section}")
            print("─" * W)

        print(f"  {s['path']}")

        for sheet, cnt in s.get("sheets", {}).items():
            print(f"    • {sheet:<32} {cnt:>8,} rows")

        dq = s.get("dq", {})
        total_rows = sum(s.get("sheets", {}).values()) or 1
        if dq:
            dirty = sum(v for k, v in dq.items() if k not in {"DUPLICATE_ID", "LATE_ARRIVING"})
            print(
                f"    DQ issues injected  (total dirty rows: {dirty:,} / {dirty/total_rows*100:.1f}%)"
            )
            for issue, cnt in dq.items():
                tag = "  [dup]" if issue == "DUPLICATE_ID" else ""
                tag = "  [late]" if issue == "LATE_ARRIVING" else tag
                pct = f"{cnt/total_rows*100:.1f}%"
                print(f"      {issue:<24} {cnt:>5}  ({pct}){tag}")

        for note in s.get("special", []):
            print(f"    ★  {note}")

    print("\n" + "═" * W + "\n")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main() -> None:
    parser = argparse.ArgumentParser(description="Generate CPG synthetic data workbooks.")
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED)
    parser.add_argument("--root", type=str, default=".", help="Repo root (default: current dir)")
    args = parser.parse_args()

    rng = np.random.default_rng(args.seed)
    root = Path(args.root)
    hist = root / "data" / "input" / "historical"
    incr = root / "data" / "input" / "incremental"
    hist.mkdir(parents=True, exist_ok=True)
    incr.mkdir(parents=True, exist_ok=True)

    promo_set = build_promo_set()
    all_dates = [HIST_START + timedelta(days=i) for i in range((HIST_END - HIST_START).days + 1)]

    stats: list[dict] = []

    # ── Reference workbook  (multi-sheet) ─────────────────────────────────────
    print("Generating reference dimensions…", flush=True)
    df_product = pd.DataFrame(
        PRODUCT_ROWS,
        columns=["sku", "category", "brand", "package_size", "list_price", "launch_date"],
    )
    df_region = pd.DataFrame(
        REGION_ROWS, columns=["region", "population", "median_income_band", "climate_zone"]
    )
    df_store = pd.DataFrame(STORE_ROWS, columns=["store_id", "region", "city", "store_type"])
    df_calendar = make_seasonal_calendar()

    ref_path = hist / "historical_data.xlsx"
    write_multisheet(
        ref_path,
        {
            "dim_product": df_product,
            "dim_region": df_region,
            "dim_store": df_store,
            "seasonal_calendar": df_calendar,
        },
    )
    stats.append(
        {
            "section": "HISTORICAL FILES",
            "path": str(ref_path.relative_to(root)),
            "sheets": {
                "dim_product": len(df_product),
                "dim_region": len(df_region),
                "dim_store": len(df_store),
                "seasonal_calendar": len(df_calendar),
            },
            "special": ["Multi-sheet workbook — exercises sheet-discovery logic"],
        }
    )

    # ── POS historical  (schema A) ────────────────────────────────────────────
    print("Generating POS history (~30 K rows, patience…)…", flush=True)
    dq_pos: dict[str, int] = {}
    df_pos = gen_pos_transactions(all_dates, promo_set, rng)
    df_pos = inject_dq(df_pos, "A", rng, dq_pos)

    # Stash last 12 transaction IDs for incremental duplicate injection
    last_pos_ids = df_pos["transaction_id"].iloc[-12:].tolist()

    pos_path = hist / "pos_sales_history.xlsx"
    write_single_sheet(pos_path, df_pos, sheet_name="Sales")
    stats.append(
        {
            "section": "HISTORICAL FILES",
            "path": str(pos_path.relative_to(root)),
            "sheets": {"Sales": len(df_pos)},
            "dq": dq_pos,
            "special": [
                "Schema A: transaction_id / ts / store_id / sku / qty / unit_price / amount / currency"
            ],
        }
    )

    # ── Online historical  (schema B, drifted) ────────────────────────────────
    print("Generating online history (~13 K rows)…", flush=True)
    dq_online: dict[str, int] = {}
    df_online = gen_online_transactions(all_dates, promo_set, rng)
    df_online = inject_dq(df_online, "B", rng, dq_online)

    last_ord_ids = df_online["order_id"].iloc[-8:].tolist()

    online_path = hist / "online_sales_history.xlsx"
    write_single_sheet(online_path, df_online, sheet_name="Orders")
    stats.append(
        {
            "section": "HISTORICAL FILES",
            "path": str(online_path.relative_to(root)),
            "sheets": {"Orders": len(df_online)},
            "dq": dq_online,
            "special": [
                "Schema B (drifted): order_id / order_datetime / location_id / product_sku / units / price_per_unit / currency",
                "No 'amount' column — revenue must be derived during ingestion",
            ],
        }
    )

    # ── Secondary feeds ────────────────────────────────────────────────────────
    for fname, df_sec, sheet in [
        ("promo_windows.xlsx", make_promo_df(), "PromoWindows"),
        ("marketing_campaigns.xlsx", make_campaign_df(), "Campaigns"),
        ("competitor_prices.xlsx", make_competitor_prices(rng), "CompetitorPrices"),
    ]:
        p = hist / fname
        write_single_sheet(p, df_sec, sheet_name=sheet)
        stats.append(
            {
                "section": "HISTORICAL FILES",
                "path": str(p.relative_to(root)),
                "sheets": {sheet: len(df_sec)},
            }
        )

    # ── Incremental batch 1: 2024-07-01  (title row + ghost sheet + dups) ─────
    print("Generating incremental batch 1…", flush=True)
    dq_b1: dict[str, int] = {}
    df_b1, dq_b1 = gen_incr_pos_batch(
        week_start=date(2024, 7, 1),
        promo_set=promo_set,
        rng=rng,
        dup_ids=last_pos_ids,  # retry-duplicate from tail of historical
        n_late=6,
        late_before=HIST_END,
    )
    b1_path = incr / "2024-07-01_pos.xlsx"
    write_single_sheet(
        b1_path,
        df_b1,
        sheet_name="Sales",
        title_row="CPG POS Export  |  batch_date=2024-07-01  |  DO NOT EDIT",
        ghost_sheet=True,
    )
    stats.append(
        {
            "section": "INCREMENTAL FILES",
            "path": str(b1_path.relative_to(root)),
            "sheets": {"Sales": len(df_b1)},
            "dq": dq_b1,
            "special": [
                "Title/metadata row ABOVE the header  (robust header-detection required)",
                "Ghost 'Notes' sheet alongside data sheet",
                f"Duplicates: {len(last_pos_ids)} transaction_ids reused from tail of pos_sales_history",
                "Late-arriving: 6 rows timestamped before 2024-06-30",
            ],
        }
    )
    # carry IDs forward for batch 3 dups
    last_b1_ids = df_b1["transaction_id"].iloc[-8:].tolist()

    # ── Incremental batch 2: 2024-07-08  (SCD2 change + late-arriving) ────────
    print("Generating incremental batch 2…", flush=True)
    df_b2, dq_b2 = gen_incr_online_batch(
        week_start=date(2024, 7, 8),
        promo_set=promo_set,
        rng=rng,
        dup_ids=last_ord_ids,  # retry-duplicate from tail of online historical
        n_late=5,
        late_before=HIST_END,
    )
    df_scd2 = make_scd2_updates()
    b2_path = incr / "2024-07-08_online.xlsx"
    write_multisheet(b2_path, {"Orders": df_b2, "product_updates": df_scd2})
    stats.append(
        {
            "section": "INCREMENTAL FILES",
            "path": str(b2_path.relative_to(root)),
            "sheets": {"Orders": len(df_b2), "product_updates": len(df_scd2)},
            "dq": dq_b2,
            "special": [
                "SCD-2 trigger: SKU001 list_price 8.99→9.49, SKU007 package_size change  (see 'product_updates' sheet)",
                f"Duplicates: {len(last_ord_ids)} order_ids reused from tail of online_sales_history",
                "Late-arriving: 5 rows timestamped before 2024-06-30",
            ],
        }
    )

    # ── Incremental batch 3: 2024-07-15  (dups from batch 1 + late rows) ──────
    print("Generating incremental batch 3…", flush=True)
    df_b3, dq_b3 = gen_incr_pos_batch(
        week_start=date(2024, 7, 15),
        promo_set=promo_set,
        rng=rng,
        dup_ids=last_b1_ids,  # retry-duplicate from batch 1
        n_late=4,
        late_before=date(2024, 7, 1),  # before batch 1 nominal date
    )
    b3_path = incr / "2024-07-15_pos.xlsx"
    write_single_sheet(b3_path, df_b3, sheet_name="Sales")
    stats.append(
        {
            "section": "INCREMENTAL FILES",
            "path": str(b3_path.relative_to(root)),
            "sheets": {"Sales": len(df_b3)},
            "dq": dq_b3,
            "special": [
                f"Duplicates: {len(last_b1_ids)} transaction_ids reused from batch 1",
                "Late-arriving: 4 rows timestamped before 2024-07-01 (pre-batch-1)",
            ],
        }
    )

    # ── Summary ───────────────────────────────────────────────────────────────
    print_summary(stats, seed=args.seed)


# ═══════════════════════════════════════════════════════════════════════════════
# LIVE WEEKLY BATCH GENERATOR  (callable from the API / UI)
# ═══════════════════════════════════════════════════════════════════════════════
def generate_weekly_batch(
    batch_type: str,
    root: Path | str = ".",
    seed: int | None = None,
) -> dict:
    """
    Generate a new incremental batch for the **current ISO week** (Monday start).

    batch_type
        ``"pos"``    → writes ``{YYYY-MM-DD}_pos.xlsx``  with a ``Sales`` sheet
                       (Schema A: transaction_id / ts / store_id / sku / qty /
                        unit_price / amount / currency)
        ``"online"`` → writes ``{YYYY-MM-DD}_online.xlsx`` with an ``Orders`` sheet
                       (Schema B: order_id / order_datetime / location_id /
                        product_sku / units / price_per_unit / currency)

    Idempotent: if the file already exists for this week, returns immediately
    with ``{"status": "exists", ...}`` and does **not** overwrite the file.

    Returns a dict with keys:
        status      "created" | "exists"
        batch_type  the type that was requested
        file        filename (not full path)
        path        absolute path to the file
        week_start  ISO date of the Monday that starts the week
        rows        number of data rows written (0 if status == "exists")
        dq          {issue_type: count} dict of injected DQ faults
    """
    global _txn_counter, _ord_counter

    if batch_type not in ("pos", "online"):
        raise ValueError(f"batch_type must be 'pos' or 'online', got {batch_type!r}")

    root_path = Path(root)
    incr_dir = root_path / "data" / "input" / "incremental"
    incr_dir.mkdir(parents=True, exist_ok=True)

    # Monday of the current week
    today = date.today()
    week_start = today - timedelta(days=today.weekday())

    fname = f"{week_start}_{batch_type}.xlsx"
    fpath = incr_dir / fname

    if fpath.exists():
        return {
            "status": "exists",
            "batch_type": batch_type,
            "file": fname,
            "path": str(fpath.resolve()),
            "week_start": str(week_start),
            "rows": 0,
            "dq": {},
        }

    # Seed: deterministic per (week_start, batch_type) so re-running the same
    # week always produces the same file — useful for debugging.
    if seed is None:
        seed = int(week_start.strftime("%Y%m%d")) + (0 if batch_type == "pos" else 1)

    rng = np.random.default_rng(seed)
    promo_set = build_promo_set()

    # Offset counters far above historical data (~35 K IDs) so generated IDs
    # never collide with the seed-42 historical workbooks.
    week_offset = int(week_start.strftime("%Y%m%d")) * 1_000
    _txn_counter = week_offset
    _ord_counter = week_offset

    if batch_type == "pos":
        df, dq = gen_incr_pos_batch(
            week_start=week_start,
            promo_set=promo_set,
            rng=rng,
            dup_ids=[],  # no explicit cross-batch dups for live batches
            n_late=3,
            late_before=week_start,
        )
        write_single_sheet(fpath, df, sheet_name="Sales")

    else:  # online
        df, dq = gen_incr_online_batch(
            week_start=week_start,
            promo_set=promo_set,
            rng=rng,
            dup_ids=[],
            n_late=2,
            late_before=week_start,
        )
        write_multisheet(fpath, {"Orders": df})

    return {
        "status": "created",
        "batch_type": batch_type,
        "file": fname,
        "path": str(fpath.resolve()),
        "week_start": str(week_start),
        "rows": len(df),
        "dq": dq,
    }


if __name__ == "__main__":
    main()
