"""
src/common/excel_io.py

Robust .xlsx reader that handles:
  • Multi-sheet workbooks  — caller picks sheets by name.
  • Ghost / stale sheets   — detected by emptiness and silently skipped.
  • Leading title row      — real header detected by max-non-null heuristic,
                             NOT assumed to be row 1.
  • Formula cells          — opened with data_only=True so cached values are used.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
from loguru import logger

# ── Ghost-sheet detection ─────────────────────────────────────────────────────


def _is_ghost(ws) -> bool:
    """True when a worksheet has no usable content."""
    if ws.max_row is None or ws.max_row == 0:
        return True
    non_empty = sum(
        1
        for row in ws.iter_rows(values_only=True)
        for cell in row
        if cell is not None and str(cell).strip() != ""
    )
    return non_empty == 0


# ── Header-row detection ──────────────────────────────────────────────────────


def _find_header_row(ws, max_scan: int = 15) -> int:
    """
    Return the 0-indexed row index of the real header row within the first
    `max_scan` rows.

    Strategy: a title/metadata row typically fills only 1-2 cells; the real
    header fills every column.  We pick the row with the most non-null cells.
    """
    best_idx, best_count = 0, 0
    limit = min(max_scan, ws.max_row or 1)
    for row_idx in range(limit):
        cells = [
            ws.cell(row=row_idx + 1, column=c).value for c in range(1, (ws.max_column or 1) + 1)
        ]
        non_null = [v for v in cells if v is not None and str(v).strip() != ""]
        if len(non_null) > best_count:
            best_count = len(non_null)
            best_idx = row_idx
    return best_idx


# ── Worksheet → DataFrame ─────────────────────────────────────────────────────


def _ws_to_df(ws, header_row: int) -> pd.DataFrame:
    """
    Convert worksheet rows starting at `header_row` (0-indexed) into a
    DataFrame.  Row at `header_row` becomes column names; everything below
    is data.
    """
    all_rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
    if not all_rows:
        return pd.DataFrame()

    raw_header = all_rows[0]
    columns = [str(c).strip() if c is not None else f"_col{i}" for i, c in enumerate(raw_header)]
    data = all_rows[1:]
    return pd.DataFrame(data, columns=columns)


# ── Public API ────────────────────────────────────────────────────────────────


def read_workbook(
    path: Path,
    *,
    target_sheets: list[str] | None = None,
) -> dict[str, pd.DataFrame]:
    """
    Open an .xlsx workbook and return ``{sheet_name: DataFrame}`` for every
    non-ghost sheet.

    Parameters
    ----------
    path
        Absolute or relative path to the workbook.
    target_sheets
        If provided, only these sheet names are returned (KeyError if absent).
        Ghost sheets within this list are still skipped with a warning.

    Notes
    -----
    * Uses ``data_only=True`` so formula cells resolve to their cached value.
    * The real header row is detected automatically — a leading metadata row
      that occupies fewer columns than the header will be ignored.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=False)
    sheets_to_process = target_sheets if target_sheets is not None else wb.sheetnames
    result: dict[str, pd.DataFrame] = {}

    for name in sheets_to_process:
        if name not in wb.sheetnames:
            raise KeyError(
                f"Sheet '{name}' not found in '{path.name}'. " f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[name]

        if _is_ghost(ws):
            logger.debug("Skipping ghost/empty sheet '{}' in '{}'", name, path.name)
            continue

        header_row = _find_header_row(ws)
        if header_row > 0:
            logger.info(
                "Title row detected in '{}!{}' — treating Excel row {} as header",
                path.name,
                name,
                header_row + 1,
            )

        df = _ws_to_df(ws, header_row)
        if df.empty:
            logger.debug(
                "Sheet '{}' in '{}' produced an empty DataFrame — skipping", name, path.name
            )
            continue

        result[name] = df
        logger.debug(
            "Read '{}!{}': {} rows × {} cols",
            path.name,
            name,
            len(df),
            len(df.columns),
        )

    wb.close()
    return result


def infer_source_system(path: Path, sheet_name: str) -> str:
    """
    Best-effort source-system label from filename + sheet name.
    Callers may override; this is used as a fallback.
    """
    stem, sheet = path.stem.lower(), sheet_name.lower()
    if "pos" in stem or sheet == "sales":
        return "POS"
    if "online" in stem or sheet == "orders":
        return "ONLINE"
    if "promo" in stem or "promo" in sheet:
        return "PROMO"
    if "campaign" in stem or "campaign" in sheet:
        return "MARKETING"
    if "competitor" in stem or "competitor" in sheet:
        return "COMPETITOR"
    return "UNKNOWN"
